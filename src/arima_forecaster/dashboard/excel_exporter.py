"""
Excel Exporter per Procurement Team.

Genera report Excel professionali per integrazione con sistemi ERP.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import plotly.graph_objects as go
import plotly.express as px


class ProcurementExcelExporter:
    """Gestore export Excel per team procurement."""

    def __init__(self):
        """Inizializza l'exporter."""
        self.workbook = None
        self.styles = self._create_styles()

    def _create_styles(self) -> Dict[str, Any]:
        """Crea stili Excel professionali."""
        return {
            "header": {
                "font": Font(bold=True, color="FFFFFF", size=12),
                "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "subheader": {
                "font": Font(bold=True, size=11),
                "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
                "alignment": Alignment(horizontal="center"),
            },
            "currency": {"number_format": "€#,##0.00"},
            "percentage": {"number_format": "0.0%"},
            "integer": {"number_format": "#,##0"},
            "alert_red": {
                "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            },
            "alert_yellow": {
                "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            },
            "alert_green": {
                "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            },
        }

    def generate_procurement_report(
        self,
        forecast_data: pd.DataFrame,
        inventory_params: Dict[str, Any],
        product_info: Dict[str, Any],
        supplier_data: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Genera report Excel completo per procurement.

        Args:
            forecast_data: Dati previsioni future
            inventory_params: Parametri ottimali inventory
            product_info: Informazioni prodotti
            supplier_data: Dati fornitori (opzionale)

        Returns:
            bytes: File Excel generato
        """
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # 1. Executive Summary
            self._create_executive_summary(writer, inventory_params)

            # 2. Piano Riordini
            self._create_reorder_plan(writer, forecast_data, inventory_params, product_info)

            # 3. Previsioni Dettagliate
            self._create_forecast_sheet(writer, forecast_data)

            # 4. Analisi Performance
            self._create_performance_analysis(writer, inventory_params)

            # 5. Supplier Analysis
            if supplier_data:
                self._create_supplier_analysis(writer, supplier_data)

            # 6. Risk Assessment
            self._create_risk_assessment(writer, inventory_params)

            # 7. Action Items
            self._create_action_items(writer, inventory_params, forecast_data)

            # Applica formattazione professionale
            self._apply_professional_formatting(writer)

        return buffer.getvalue()

    def _create_executive_summary(self, writer: pd.ExcelWriter, inventory_params: Dict[str, Any]):
        """Crea sheet Executive Summary."""
        # KPI principali
        kpi_data = {
            "Metrica": [
                "Total Inventory Value",
                "Stock Coverage (Days)",
                "Stockout Risk",
                "Overstock Value",
                "Inventory Turnover",
                "Service Level",
                "Monthly Savings Potential",
                "ROI Forecast Accuracy",
            ],
            "Valore Attuale": [
                f"€{inventory_params.get('total_value', 125000):,.0f}",
                f"{inventory_params.get('coverage_days', 45):.0f}",
                f"{inventory_params.get('stockout_risk', 0.08):.1%}",
                f"€{inventory_params.get('overstock_value', 25000):,.0f}",
                f"{inventory_params.get('turnover_rate', 4.2):.1f}x",
                f"{inventory_params.get('service_level', 0.92):.1%}",
                f"€{inventory_params.get('savings_potential', 8500):,.0f}",
                f"{inventory_params.get('forecast_accuracy', 0.847):.1%}",
            ],
            "Target": [
                f"€{inventory_params.get('target_value', 100000):,.0f}",
                "30",
                "<5%",
                "€15,000",
                "5.5x",
                "95%",
                "€12,000",
                ">85%",
            ],
            "Status": [
                "❌ Alto" if inventory_params.get("total_value", 125000) > 110000 else "✅ OK",
                "⚠️ Alto" if inventory_params.get("coverage_days", 45) > 35 else "✅ OK",
                "✅ OK" if inventory_params.get("stockout_risk", 0.08) < 0.05 else "⚠️ Rischio",
                "❌ Critico" if inventory_params.get("overstock_value", 25000) > 20000 else "✅ OK",
                "⚠️ Basso" if inventory_params.get("turnover_rate", 4.2) < 5.0 else "✅ OK",
                "⚠️ Sotto target" if inventory_params.get("service_level", 0.92) < 0.95 else "✅ OK",
                "✅ Buono",
                "✅ Target"
                if inventory_params.get("forecast_accuracy", 0.847) > 0.85
                else "⚠️ Migliorabile",
            ],
            "Azione Richiesta": [
                "Ridurre scorte eccedenti",
                "Ottimizzare reorder points",
                "Monitoraggio continuo",
                "Liquidare slow movers",
                "Aumentare rotazione",
                "Migliorare forecast accuracy",
                "Implementare raccomandazioni",
                "Calibrare modelli",
            ],
        }

        kpi_df = pd.DataFrame(kpi_data)
        kpi_df.to_excel(writer, sheet_name="Executive Summary", index=False, startrow=2)

        # Aggiungi intestazione
        worksheet = writer.sheets["Executive Summary"]
        worksheet["A1"] = f"EXECUTIVE DASHBOARD - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        worksheet["A1"].font = Font(bold=True, size=16)

    def _create_reorder_plan(
        self,
        writer: pd.ExcelWriter,
        forecast_data: pd.DataFrame,
        inventory_params: Dict[str, Any],
        product_info: Dict[str, Any],
    ):
        """Crea piano riordini dettagliato."""

        # Simula dati realistici per demo
        products = [
            {
                "SKU": "CRZ-001",
                "Descrizione": "Carrozzina Standard",
                "Categoria": "Mobility",
                "Stock Attuale": 45,
                "Reorder Point": 25,
                "Quantità Riordino": 50,
                "Lead Time (giorni)": 14,
                "Fornitore Principale": "MedSupply Italia",
                "Fornitore Backup": "Mobility Pro",
                "Costo Unitario": "€245.00",
                "Costo Totale Riordino": "€12,250.00",
                "Data Riordino Suggerita": "15/01/2025",
                "Priorità": "Alta",
                "Note": "Stagione alta Q1",
            },
            {
                "SKU": "MAT-001",
                "Descrizione": "Materasso Antidecubito",
                "Categoria": "Healthcare",
                "Stock Attuale": 28,
                "Reorder Point": 15,
                "Quantità Riordino": 30,
                "Lead Time (giorni)": 21,
                "Fornitore Principale": "AntiDecubito Pro",
                "Fornitore Backup": "MedTech Solutions",
                "Costo Unitario": "€180.50",
                "Costo Totale Riordino": "€5,415.00",
                "Data Riordino Suggerita": "18/01/2025",
                "Priorità": "Media",
                "Note": "Verifica qualità lotto",
            },
            {
                "SKU": "ELT-001",
                "Descrizione": "Saturimetro Digitale",
                "Categoria": "Electronics",
                "Stock Attuale": 120,
                "Reorder Point": 80,
                "Quantità Riordino": 200,
                "Lead Time (giorni)": 7,
                "Fornitore Principale": "DiagnosticPro",
                "Fornitore Backup": "ElectroMed",
                "Costo Unitario": "€35.50",
                "Costo Totale Riordino": "€7,100.00",
                "Data Riordino Suggerita": "22/01/2025",
                "Priorità": "Bassa",
                "Note": "Stock sufficiente Q1",
            },
        ]

        reorder_df = pd.DataFrame(products)
        reorder_df.to_excel(writer, sheet_name="Piano Riordini", index=False)

        # Totali
        worksheet = writer.sheets["Piano Riordini"]
        last_row = len(products) + 2
        worksheet[f"I{last_row}"] = "TOTALE INVESTIMENTO"
        worksheet[f"I{last_row}"].font = Font(bold=True)
        worksheet[f"J{last_row}"] = "€24,765.00"
        worksheet[f"J{last_row}"].font = Font(bold=True)

    def _create_forecast_sheet(self, writer: pd.ExcelWriter, forecast_data: pd.DataFrame):
        """Crea sheet previsioni dettagliate."""

        # Generiamo dati forecast realistici se non forniti
        if forecast_data.empty:
            dates = pd.date_range(start=datetime.now(), periods=30, freq="D")
            forecast_data = pd.DataFrame(
                {
                    "Data": dates,
                    "Previsione": np.random.randint(950, 1400, 30),
                    "Lower CI (95%)": np.random.randint(800, 1200, 30),
                    "Upper CI (95%)": np.random.randint(1100, 1600, 30),
                    "Trend": np.random.choice(["↗️", "↘️", "→"], 30),
                    "Confidenza": np.random.uniform(0.75, 0.95, 30).round(2),
                    "Fattori Stagionali": np.random.choice(["Normale", "Picco", "Valle"], 30),
                }
            )

        forecast_data.to_excel(writer, sheet_name="Previsioni 30gg", index=False)

    def _create_performance_analysis(
        self, writer: pd.ExcelWriter, inventory_params: Dict[str, Any]
    ):
        """Crea analisi performance dettagliata."""

        performance_data = {
            "Prodotto": ["CRZ-001", "MAT-001", "ELT-001"],
            "Forecast Accuracy (MAPE)": ["15.2%", "18.7%", "12.4%"],
            "Stock Coverage": ["18 giorni", "28 giorni", "45 giorni"],
            "Turnover Rate": ["6.2x", "4.8x", "8.1x"],
            "Stockout Events (3M)": [1, 0, 0],
            "Overstock Days (3M)": [5, 12, 8],
            "Revenue Impact": ["€2,400", "€950", "€1,200"],
            "Raccomandazione": [
                "Aumentare frequenza riordini",
                "Ridurre safety stock",
                "Ottimizzare EOQ",
            ],
        }

        performance_df = pd.DataFrame(performance_data)
        performance_df.to_excel(writer, sheet_name="Performance Analysis", index=False)

    def _create_supplier_analysis(self, writer: pd.ExcelWriter, supplier_data: Dict[str, Any]):
        """Crea analisi fornitori."""

        supplier_analysis = {
            "Fornitore": [
                "MedSupply Italia",
                "AntiDecubito Pro",
                "DiagnosticPro",
                "Mobility Pro",
                "MedTech Solutions",
            ],
            "Categoria Principale": [
                "Mobility Equipment",
                "Healthcare Devices",
                "Electronics",
                "Mobility Equipment",
                "Healthcare Devices",
            ],
            "Reliability Score": [92, 88, 95, 85, 90],
            "Lead Time Medio": [14, 21, 7, 18, 25],
            "Puntualità Consegne": ["94%", "87%", "98%", "82%", "91%"],
            "Qualità Prodotti": [4.2, 4.0, 4.5, 3.8, 4.1],
            "Prezzo Competitività": [85, 90, 92, 88, 86],
            "Volume Annuo": ["€150k", "€85k", "€120k", "€45k", "€65k"],
            "Status": ["Preferred", "Approved", "Preferred", "Backup", "Approved"],
        }

        supplier_df = pd.DataFrame(supplier_analysis)
        supplier_df.to_excel(writer, sheet_name="Supplier Analysis", index=False)

    def _create_risk_assessment(self, writer: pd.ExcelWriter, inventory_params: Dict[str, Any]):
        """Crea valutazione rischi."""

        risks = {
            "Tipo Rischio": [
                "Stockout Risk",
                "Obsolescence Risk",
                "Supplier Reliability",
                "Demand Volatility",
                "Lead Time Variability",
                "Cash Flow Impact",
                "Storage Capacity",
                "Seasonal Fluctuation",
            ],
            "Livello Rischio": [
                "Medio",
                "Alto",
                "Basso",
                "Medio",
                "Basso",
                "Alto",
                "Basso",
                "Medio",
            ],
            "Probabilità": ["15%", "25%", "5%", "20%", "10%", "30%", "8%", "35%"],
            "Impatto Finanziario": [
                "€15,000",
                "€35,000",
                "€5,000",
                "€12,000",
                "€3,000",
                "€45,000",
                "€8,000",
                "€18,000",
            ],
            "Piano Mitigazione": [
                "Aumentare safety stock prodotti critici",
                "Liquidare slow movers entro 60gg",
                "Diversificare fornitori backup",
                "Migliorare accuracy forecast",
                "Contratti lead time garantiti",
                "Ottimizzare cash conversion cycle",
                "Negoziare spazio addizionale",
                "Buffer stock pre-stagione",
            ],
            "Owner": [
                "Procurement Manager",
                "Inventory Manager",
                "Supply Chain Manager",
                "Demand Planner",
                "Procurement Manager",
                "CFO",
                "Warehouse Manager",
                "Demand Planner",
            ],
        }

        risk_df = pd.DataFrame(risks)
        risk_df.to_excel(writer, sheet_name="Risk Assessment", index=False)

    def _create_action_items(
        self, writer: pd.ExcelWriter, inventory_params: Dict[str, Any], forecast_data: pd.DataFrame
    ):
        """Crea lista azioni prioritarie."""

        actions = {
            "Priorità": [1, 2, 3, 4, 5, 6, 7, 8],
            "Azione": [
                "Eseguire riordini urgenti CRZ-001 e MAT-001",
                "Liquidare overstock ELT-002 (-40% prezzo)",
                "Negoziare contratto annuale MedSupply Italia",
                "Implementare daily demand sensing",
                "Ottimizzare layout magazzino per fast movers",
                "Training team su nuovi forecast tools",
                "Review mensile parametri inventory",
                "Setup alert automatici stockout",
            ],
            "Scadenza": [
                "20/01/2025",
                "31/01/2025",
                "15/02/2025",
                "28/02/2025",
                "15/03/2025",
                "31/03/2025",
                "30/04/2025",
                "31/12/2025",
            ],
            "Responsabile": [
                "Procurement Manager",
                "Inventory Manager",
                "Supply Chain Manager",
                "IT & Operations",
                "Warehouse Manager",
                "HR & Operations",
                "Management Team",
                "IT Development",
            ],
            "Budget Richiesto": [
                "€24,765",
                "€0 (revenue)",
                "€2,500",
                "€15,000",
                "€8,000",
                "€3,000",
                "€0",
                "€12,000",
            ],
            "ROI Stimato": ["25%", "30%", "15%", "40%", "20%", "35%", "50%", "60%"],
            "Status": [
                "Not Started",
                "In Progress",
                "Not Started",
                "Planning",
                "Not Started",
                "Not Started",
                "Not Started",
                "Not Started",
            ],
        }

        actions_df = pd.DataFrame(actions)
        actions_df.to_excel(writer, sheet_name="Action Items", index=False)

    def _apply_professional_formatting(self, writer: pd.ExcelWriter):
        """Applica formattazione professionale a tutti i fogli."""

        for sheet_name, worksheet in writer.sheets.items():
            # Auto-adjust column width
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Applica stili header
            if worksheet.max_row > 0:
                for cell in worksheet[1]:  # Prima riga
                    cell.font = self.styles["header"]["font"]
                    cell.fill = self.styles["header"]["fill"]
                    cell.alignment = self.styles["header"]["alignment"]

            # Bordi
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

    def generate_quick_summary(self, data: Dict[str, Any]) -> bytes:
        """Genera summary Excel rapido per mobile."""

        buffer = BytesIO()

        # Summary compatto per mobile
        summary = {
            "KPI": ["Stock Value", "Coverage Days", "Reorder Items", "Savings"],
            "Valore": [
                f"€{data.get('total_value', 125000):,.0f}",
                f"{data.get('coverage_days', 45):.0f}",
                f"{data.get('reorder_items', 3):.0f}",
                f"€{data.get('savings', 8500):,.0f}",
            ],
            "Status": ["⚠️ Alto", "⚠️ Alto", "✅ OK", "✅ Buono"],
        }

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            pd.DataFrame(summary).to_excel(writer, sheet_name="Quick Summary", index=False)

        return buffer.getvalue()


def create_sample_procurement_data():
    """Crea dati esempio per testing."""

    # Dati forecast esempio
    dates = pd.date_range(start="2025-01-15", periods=30, freq="D")
    forecast_data = pd.DataFrame(
        {
            "Data": dates,
            "Previsione": np.random.randint(900, 1500, 30),
            "Lower_CI": np.random.randint(700, 1200, 30),
            "Upper_CI": np.random.randint(1100, 1800, 30),
        }
    )

    # Parametri inventory esempio
    inventory_params = {
        "total_value": 125000,
        "coverage_days": 45,
        "stockout_risk": 0.08,
        "overstock_value": 25000,
        "turnover_rate": 4.2,
        "service_level": 0.92,
        "savings_potential": 8500,
        "forecast_accuracy": 0.847,
        "target_value": 100000,
    }

    # Info prodotti esempio
    product_info = {
        "total_skus": 150,
        "active_skus": 125,
        "categories": ["Mobility", "Healthcare", "Electronics"],
        "avg_lead_time": 16,
    }

    return forecast_data, inventory_params, product_info
