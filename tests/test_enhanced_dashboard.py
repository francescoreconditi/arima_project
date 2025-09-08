"""
Test per Dashboard Evolution - Mobile, Excel Export, What-If Simulator.

Test integrati per tutte le nuove funzionalità enterprise.
"""

import pytest
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO
import openpyxl

from arima_forecaster.dashboard.mobile_responsive import (
    MobileResponsiveManager,
    init_responsive_dashboard
)
from arima_forecaster.dashboard.excel_exporter import (
    ProcurementExcelExporter,
    create_sample_procurement_data
)
from arima_forecaster.dashboard.scenario_simulator import (
    WhatIfScenarioSimulator,
    ScenarioParameters,
    ScenarioType,
    create_sample_base_data
)


class TestMobileResponsiveManager:
    """Test per gestore mobile responsive."""
    
    @pytest.fixture
    def responsive_manager(self):
        """Crea instance manager responsive."""
        return MobileResponsiveManager()
    
    def test_device_detection(self, responsive_manager):
        """Test rilevamento tipo dispositivo."""
        # Test con diverse larghezze schermo
        assert isinstance(responsive_manager.is_mobile, bool)
        assert isinstance(responsive_manager.is_tablet, bool)
        
        # Test mutua esclusività
        if responsive_manager.is_mobile:
            assert not responsive_manager.is_tablet
    
    def test_layout_config_generation(self, responsive_manager):
        """Test generazione configurazione layout."""
        config = responsive_manager.get_layout_config()
        
        required_keys = ['columns', 'chart_height', 'sidebar_state', 'font_size', 'spacing']
        for key in required_keys:
            assert key in config
        
        # Test valori ragionevoli
        assert 1 <= config['columns'] <= 3
        assert 200 <= config['chart_height'] <= 600
        assert config['sidebar_state'] in ['collapsed', 'expanded', 'auto']
    
    def test_responsive_css_generation(self, responsive_manager):
        """Test generazione CSS responsive."""
        # Dovrebbe essere chiamabile senza errori
        try:
            responsive_manager.apply_responsive_css()
            assert True
        except Exception as e:
            pytest.fail(f"CSS generation failed: {e}")
    
    def test_responsive_metrics_display(self, responsive_manager):
        """Test visualizzazione metriche responsive."""
        sample_metrics = {
            'Revenue': {'value': '€125,000', 'delta': '+15%'},
            'Orders': {'value': '234', 'delta': '+5'},
            'Efficiency': {'value': '92%', 'delta': '+2.1%'}
        }
        
        # Should not raise errors
        try:
            # Simula chiamata (senza Streamlit reale)
            config = responsive_manager.get_layout_config()
            assert len(sample_metrics) >= config['columns'] or config['columns'] == 1
        except Exception as e:
            pytest.fail(f"Metrics display failed: {e}")


class TestProcurementExcelExporter:
    """Test per Excel exporter procurement."""
    
    @pytest.fixture
    def excel_exporter(self):
        """Crea instance exporter."""
        return ProcurementExcelExporter()
    
    @pytest.fixture
    def sample_data(self):
        """Crea dati esempio per test."""
        return create_sample_procurement_data()
    
    def test_excel_styles_creation(self, excel_exporter):
        """Test creazione stili Excel."""
        styles = excel_exporter.styles
        
        required_styles = ['header', 'subheader', 'currency', 'percentage']
        for style in required_styles:
            assert style in styles
        
        # Test header style structure
        header = styles['header']
        assert 'font' in header or isinstance(header, dict)
    
    def test_sample_data_generation(self, sample_data):
        """Test generazione dati sample."""
        forecast_data, inventory_params, product_info = sample_data
        
        # Test forecast data
        assert isinstance(forecast_data, pd.DataFrame)
        assert len(forecast_data) > 0
        assert 'Data' in forecast_data.columns or len(forecast_data.columns) >= 2
        
        # Test inventory params
        assert isinstance(inventory_params, dict)
        required_params = ['total_value', 'coverage_days', 'stockout_risk']
        for param in required_params:
            assert param in inventory_params
            assert isinstance(inventory_params[param], (int, float))
        
        # Test product info
        assert isinstance(product_info, dict)
        assert 'total_skus' in product_info
    
    def test_procurement_report_generation(self, excel_exporter, sample_data):
        """Test generazione report procurement completo."""
        forecast_data, inventory_params, product_info = sample_data
        
        try:
            excel_bytes = excel_exporter.generate_procurement_report(
                forecast_data=forecast_data,
                inventory_params=inventory_params,
                product_info=product_info
            )
            
            # Verifica che sia un file Excel valido
            assert isinstance(excel_bytes, bytes)
            assert len(excel_bytes) > 1000  # File ragionevole
            
            # Test lettura Excel
            buffer = BytesIO(excel_bytes)
            workbook = openpyxl.load_workbook(buffer)
            
            # Verifica sheet esistenti
            expected_sheets = ['Executive Summary', 'Piano Riordini', 'Previsioni 30gg']
            for sheet in expected_sheets:
                assert sheet in workbook.sheetnames
            
            # Verifica contenuto Executive Summary
            exec_sheet = workbook['Executive Summary']
            assert exec_sheet.max_row > 1  # Almeno header + dati
            assert exec_sheet.max_column >= 4  # Almeno 4 colonne
            
        except Exception as e:
            pytest.fail(f"Procurement report generation failed: {e}")
    
    def test_quick_summary_generation(self, excel_exporter):
        """Test generazione summary rapido."""
        sample_data = {
            'total_value': 125000,
            'coverage_days': 45,
            'reorder_items': 3,
            'savings': 8500
        }
        
        try:
            excel_bytes = excel_exporter.generate_quick_summary(sample_data)
            
            # Verifica dimensioni ragionevoli
            assert isinstance(excel_bytes, bytes)
            assert 500 <= len(excel_bytes) <= 50000
            
            # Test lettura
            buffer = BytesIO(excel_bytes)
            workbook = openpyxl.load_workbook(buffer)
            assert 'Quick Summary' in workbook.sheetnames
            
        except Exception as e:
            pytest.fail(f"Quick summary generation failed: {e}")
    
    def test_professional_formatting(self, excel_exporter, sample_data):
        """Test formattazione professionale Excel."""
        forecast_data, inventory_params, product_info = sample_data
        
        excel_bytes = excel_exporter.generate_procurement_report(
            forecast_data, inventory_params, product_info
        )
        
        buffer = BytesIO(excel_bytes)
        workbook = openpyxl.load_workbook(buffer)
        
        # Test formattazione header
        exec_sheet = workbook['Executive Summary']
        if exec_sheet.max_row > 2:  # Se ci sono dati
            # Prima riga dovrebbe essere formattata
            first_cell = exec_sheet.cell(1, 1)
            # Verifica che abbia qualche formattazione
            has_formatting = (
                first_cell.font.bold is not None or
                first_cell.fill.start_color.index != '00000000' or
                first_cell.alignment.horizontal is not None
            )
            assert has_formatting, "Header should have formatting"


class TestWhatIfScenarioSimulator:
    """Test per What-If Scenario Simulator."""
    
    @pytest.fixture
    def scenario_simulator(self):
        """Crea instance simulator."""
        return WhatIfScenarioSimulator()
    
    @pytest.fixture
    def sample_scenario_data(self):
        """Crea dati scenario di test."""
        return create_sample_base_data()
    
    def test_scenario_parameters_creation(self, scenario_simulator):
        """Test creazione parametri scenario."""
        # Test parametri default
        params = ScenarioParameters()
        
        # Verifica valori default ragionevoli
        assert 0 <= params.marketing_boost <= 500
        assert -50 <= params.price_change <= 50
        assert 0.1 <= params.seasonality_factor <= 5.0
        assert 50 <= params.supplier_reliability <= 100
        assert -30 <= params.lead_time_change <= 30
    
    def test_predefined_scenarios(self, scenario_simulator):
        """Test scenari predefiniti."""
        # Test tutti gli scenari predefiniti
        for scenario_type in ScenarioType:
            if scenario_type != ScenarioType.CUSTOM:
                params = scenario_simulator._get_predefined_scenario(scenario_type)
                assert isinstance(params, ScenarioParameters)
                
                # Verifica che i parametri siano ragionevoli
                assert -100 <= params.marketing_boost <= 500
                assert -50 <= params.price_change <= 50
                assert 0.1 <= params.seasonality_factor <= 5.0
    
    def test_demand_impact_calculation(self, scenario_simulator):
        """Test calcolo impatto sulla domanda."""
        # Scenario positivo
        positive_params = ScenarioParameters(
            marketing_boost=50.0,
            price_change=-10.0,
            seasonality_factor=1.2
        )
        
        demand_multiplier = scenario_simulator._calculate_demand_impact(positive_params)
        assert demand_multiplier > 1.0, "Positive scenario should increase demand"
        
        # Scenario negativo
        negative_params = ScenarioParameters(
            marketing_boost=-30.0,
            price_change=20.0,
            competitor_impact=-25.0
        )
        
        demand_multiplier = scenario_simulator._calculate_demand_impact(negative_params)
        assert 0.1 <= demand_multiplier <= 1.5, "Demand multiplier should be reasonable"
    
    def test_supply_constraints_calculation(self, scenario_simulator):
        """Test calcolo vincoli supply chain."""
        # Vincoli severi
        constrained_params = ScenarioParameters(
            supplier_reliability=70.0,
            lead_time_change=14,
            capacity_limit=60.0
        )
        
        supply_factor = scenario_simulator._calculate_supply_constraints(constrained_params)
        assert 0.1 <= supply_factor <= 1.0, "Supply constraints should reduce capacity"
        
        # Condizioni ideali
        ideal_params = ScenarioParameters(
            supplier_reliability=100.0,
            lead_time_change=-5,
            capacity_limit=120.0
        )
        
        supply_factor = scenario_simulator._calculate_supply_constraints(ideal_params)
        assert supply_factor >= 0.9, "Ideal conditions should have minimal impact"
    
    def test_economic_impact_calculation(self, scenario_simulator):
        """Test calcolo impatto economico."""
        # Alta inflazione
        high_inflation = ScenarioParameters(
            inflation_rate=10.0,
            exchange_rate=15.0,
            interest_rate=8.0
        )
        
        cost_multiplier = scenario_simulator._calculate_economic_impact(high_inflation)
        assert cost_multiplier > 1.0, "High inflation should increase costs"
        
        # Condizioni stabili
        stable_conditions = ScenarioParameters(
            inflation_rate=2.0,
            exchange_rate=0.0,
            interest_rate=3.0
        )
        
        cost_multiplier = scenario_simulator._calculate_economic_impact(stable_conditions)
        assert 0.9 <= cost_multiplier <= 1.2, "Stable conditions should have moderate impact"
    
    def test_full_scenario_simulation(self, scenario_simulator, sample_scenario_data):
        """Test simulazione scenario completa."""
        base_forecast, base_metrics = sample_scenario_data
        
        # Test scenario marketing boost
        marketing_scenario = ScenarioParameters(
            marketing_boost=100.0,
            price_change=-15.0
        )
        
        scenario_forecast, results = scenario_simulator.run_scenario_simulation(
            marketing_scenario, base_forecast, base_metrics
        )
        
        # Verifica output
        assert isinstance(scenario_forecast, pd.Series)
        assert len(scenario_forecast) == len(base_forecast)
        
        # Verifica risultati
        assert hasattr(results, 'revenue_impact')
        assert hasattr(results, 'inventory_investment')
        assert hasattr(results, 'service_level')
        assert hasattr(results, 'recommendations')
        
        # Verifica valori ragionevoli
        assert isinstance(results.revenue_impact, (int, float))
        assert isinstance(results.service_level, (int, float))
        assert 70 <= results.service_level <= 100
        assert isinstance(results.recommendations, list)
        assert len(results.recommendations) > 0
    
    def test_scenario_visualization_creation(self, scenario_simulator, sample_scenario_data):
        """Test creazione visualizzazioni scenario."""
        base_forecast, base_metrics = sample_scenario_data
        
        # Crea scenario semplice
        params = ScenarioParameters(marketing_boost=50.0)
        scenario_forecast, results = scenario_simulator.run_scenario_simulation(
            params, base_forecast, base_metrics
        )
        
        # Test creazione grafico
        try:
            fig = scenario_simulator.create_scenario_visualization(
                base_forecast, scenario_forecast, results
            )
            
            # Verifica che sia un oggetto Plotly valido
            assert hasattr(fig, 'data')
            assert hasattr(fig, 'layout')
            assert len(fig.data) > 0
            
        except Exception as e:
            pytest.fail(f"Visualization creation failed: {e}")
    
    def test_recommendations_generation(self, scenario_simulator):
        """Test generazione raccomandazioni."""
        # Test diversi tipi di scenario
        scenarios_to_test = [
            (50.0, 2.0),   # Scenario positivo
            (-30.0, 8.0),  # Scenario negativo
            (5.0, 1.0)     # Scenario neutro
        ]
        
        for revenue_change, service_risk in scenarios_to_test:
            params = ScenarioParameters()
            recommendations = scenario_simulator._generate_recommendations(
                params, revenue_change, service_risk
            )
            
            assert isinstance(recommendations, list)
            assert len(recommendations) > 0
            assert all(isinstance(rec, str) for rec in recommendations)
            assert all(len(rec) > 10 for rec in recommendations)  # Raccomandazioni meaningful


class TestDashboardIntegration:
    """Test integrazione dashboard completa."""
    
    def test_component_initialization(self):
        """Test inizializzazione componenti dashboard."""
        try:
            # Test inizializzazione responsive manager
            responsive = MobileResponsiveManager()
            assert responsive is not None
            
            # Test inizializzazione exporter
            exporter = ProcurementExcelExporter()
            assert exporter is not None
            
            # Test inizializzazione simulator
            simulator = WhatIfScenarioSimulator()
            assert simulator is not None
            
        except Exception as e:
            pytest.fail(f"Component initialization failed: {e}")
    
    def test_data_flow_integration(self):
        """Test flusso dati tra componenti."""
        # Crea dati di test
        base_forecast, base_metrics = create_sample_base_data()
        forecast_data, inventory_params, product_info = create_sample_procurement_data()
        
        # Test scenario simulation → Excel export
        simulator = WhatIfScenarioSimulator()
        exporter = ProcurementExcelExporter()
        
        # Run scenario
        params = ScenarioParameters(marketing_boost=25.0)
        scenario_forecast, results = simulator.run_scenario_simulation(
            params, base_forecast, base_metrics
        )
        
        # Export results
        try:
            excel_data = exporter.generate_procurement_report(
                forecast_data, inventory_params, product_info
            )
            assert len(excel_data) > 1000
            
        except Exception as e:
            pytest.fail(f"Data flow integration failed: {e}")
    
    def test_mobile_excel_integration(self):
        """Test integrazione mobile con export Excel."""
        responsive = MobileResponsiveManager()
        exporter = ProcurementExcelExporter()
        
        # Test quick export per mobile
        sample_data = {'total_value': 50000, 'coverage_days': 30}
        
        if responsive.is_mobile:
            # Mobile dovrebbe usare quick summary
            excel_data = exporter.generate_quick_summary(sample_data)
        else:
            # Desktop può usare report completo
            forecast_data, inventory_params, product_info = create_sample_procurement_data()
            excel_data = exporter.generate_procurement_report(
                forecast_data, inventory_params, product_info
            )
        
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 100
    
    @pytest.mark.integration
    def test_complete_workflow(self):
        """Test workflow completo dashboard."""
        # Simula workflow completo utente
        
        # 1. Inizializza componenti
        responsive = MobileResponsiveManager()
        simulator = WhatIfScenarioSimulator()
        exporter = ProcurementExcelExporter()
        
        # 2. Ottieni dati base
        base_forecast, base_metrics = create_sample_base_data()
        
        # 3. Configura scenario
        scenario_params = ScenarioParameters(
            marketing_boost=75.0,
            price_change=-20.0,
            seasonality_factor=1.5
        )
        
        # 4. Esegui simulazione
        scenario_forecast, results = simulator.run_scenario_simulation(
            scenario_params, base_forecast, base_metrics
        )
        
        # 5. Genera visualizzazione
        fig = simulator.create_scenario_visualization(
            base_forecast, scenario_forecast, results
        )
        
        # 6. Adatta per mobile se necessario
        if responsive.is_mobile:
            fig = responsive.create_responsive_chart(fig, "Scenario Results")
        
        # 7. Esporta risultati
        forecast_data, inventory_params, product_info = create_sample_procurement_data()
        
        # Aggiorna parametri con risultati scenario
        inventory_params['revenue_impact'] = results.revenue_impact
        inventory_params['service_level'] = results.service_level
        
        excel_report = exporter.generate_procurement_report(
            forecast_data, inventory_params, product_info
        )
        
        # Verifica workflow completato
        assert len(scenario_forecast) > 0
        assert results.revenue_impact != 0
        assert len(excel_report) > 2000
        assert hasattr(fig, 'data')
        
    def test_error_handling(self):
        """Test gestione errori robusta."""
        simulator = WhatIfScenarioSimulator()
        exporter = ProcurementExcelExporter()
        
        # Test con dati invalidi
        invalid_forecast = pd.Series([])  # Serie vuota
        invalid_metrics = {}  # Metriche vuote
        
        try:
            # Dovrebbe gestire gracefully
            params = ScenarioParameters()
            scenario_forecast, results = simulator.run_scenario_simulation(
                params, invalid_forecast, invalid_metrics
            )
            
            # Risultati dovrebbero essere ragionevoli anche con input invalidi
            assert isinstance(scenario_forecast, pd.Series)
            assert hasattr(results, 'recommendations')
            
        except Exception as e:
            # Se fallisce, almeno non dovrebbe crashare
            assert "division by zero" not in str(e).lower()
    
    def test_performance_benchmarks(self):
        """Test benchmark performance componenti."""
        import time
        
        # Test performance Excel export
        start = time.time()
        exporter = ProcurementExcelExporter()
        forecast_data, inventory_params, product_info = create_sample_procurement_data()
        
        excel_data = exporter.generate_procurement_report(
            forecast_data, inventory_params, product_info
        )
        excel_time = time.time() - start
        
        # Export dovrebbe completare in <5 secondi
        assert excel_time < 5.0, f"Excel export too slow: {excel_time:.2f}s"
        
        # Test performance scenario simulation
        start = time.time()
        simulator = WhatIfScenarioSimulator()
        base_forecast, base_metrics = create_sample_base_data()
        
        params = ScenarioParameters(marketing_boost=100.0)
        scenario_forecast, results = simulator.run_scenario_simulation(
            params, base_forecast, base_metrics
        )
        scenario_time = time.time() - start
        
        # Simulazione dovrebbe completare in <2 secondi
        assert scenario_time < 2.0, f"Scenario simulation too slow: {scenario_time:.2f}s"